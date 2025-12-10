from __future__ import annotations
from typing import List, Any

from datetime import datetime, timedelta

import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from . import config

DEFAULT_FONT = Font(name="Times New Roman", size=13)


# ── Google Sheets ────────────────────────────────────────────────────────────

def open_sheet(title: str):
    
    # Authenticate and open the given worksheet by title.
    # Returns a gspread.models.Worksheet object.
    
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(
        config.CREDS_FILE, scope
    )
    client = gspread.authorize(creds)
    sh = client.open_by_key(config.SHEET_ID)
    return sh.worksheet(title)


def clear_range(ws, cell_range: str) -> None:
    #Clear a rectangular range in a worksheet (e.g., 'B11:L1000').
    ws.batch_clear([cell_range])


def auto_stamp_gsheet(ws, tz_offset_hours: int = 0) -> None:
    # Stamp the date and time the program was run in B1 ("Last Updated:"), C1 (date), D1 (time).
    now = datetime.now() + timedelta(hours=tz_offset_hours)
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")

    ws.update(
        range_name="B1:D1",
        values=[["Last Updated:", date_str, time_str]],
        value_input_option="USER_ENTERED",
    )


# ── Excel workbook helpers (optional; use if you still generate XLSX) ───────

def auto_stamp_excel(ws) -> None:
    """Stamp date/time in B1–D1 in a local Excel sheet."""
    now = datetime.now()
    ws["B1"].value = "Last Updated:"
    ws["C1"].value = now
    ws["D1"].value = now

    ws["C1"].number_format = r'dd"::"mmmm"::"yyyy'
    ws["D1"].number_format = r'hh:mm'

    for cell_id in ("B1", "C1", "D1"):
        ws[cell_id].font = DEFAULT_FONT
